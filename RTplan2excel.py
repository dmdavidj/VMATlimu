import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import pydicom
import openpyxl

def load_folder():
    folder_selected = filedialog.askdirectory()
    if folder_selected:
        folder_path.set(folder_selected)
        process_files(folder_selected)

def process_files(folder):
    for filename in os.listdir(folder):
        if filename.lower().endswith(".dcm"):
            filepath = os.path.join(folder, filename)
            try:
                dicom_data = pydicom.dcmread(filepath)
                if 'RTPlanName' not in dicom_data or 'BeamSequence' not in dicom_data:
                    continue
                
                plan_name = dicom_data.RTPlanName
                beam_set_name = dicom_data.BeamSequence[0].TreatmentMachineName
                beam_name = dicom_data.BeamSequence[0].BeamName

                workbook = openpyxl.Workbook()
                sheet = workbook.active

                sheet.cell(row=1, column=1, value="Plan Name")
                sheet.cell(row=1, column=2, value=plan_name)
                sheet.cell(row=2, column=1, value="Beam Set Name (300A 0002)")
                sheet.cell(row=2, column=2, value=beam_set_name)
                sheet.cell(row=3, column=1, value="Beam Name")
                sheet.cell(row=3, column=2, value=beam_name)
                sheet.cell(row=4, column=1, value="Segment ID")
                sheet.cell(row=4, column=2, value="Gantry Angle")

                leaf_count = 80  # Assuming there are always 80 leaves

                for i in range(leaf_count):
                    sheet.cell(row=5, column=i+3, value=f"Leaf 0 MLC {i+1}")
                    sheet.cell(row=5, column=leaf_count+i+3, value=f"Leaf 1 MLC {i+1}")

                current_row = 6

                for beam in dicom_data.BeamSequence:
                    for cp_idx, control_point in enumerate(beam.ControlPointSequence):
                        if 'BeamLimitingDevicePositionSequence' in control_point:
                            for device_position in control_point.BeamLimitingDevicePositionSequence:
                                leaf_positions = device_position.LeafJawPositions
                                # Process and store the leaf positions
                                if len(leaf_positions) >= leaf_count * 2:
                                    sheet.cell(row=current_row, column=1, value=cp_idx + 1)
                                    gantry_angle = control_point.GantryAngle if 'GantryAngle' in control_point else 'N/A'
                                    sheet.cell(row=current_row, column=2, value=gantry_angle)
                                    for i in range(leaf_count):
                                        sheet.cell(row=current_row, column=i+3, value=leaf_positions[i])
                                        sheet.cell(row=current_row, column=leaf_count+i+3, value=leaf_positions[leaf_count + i])
                                    current_row += 1

                excel_filename = f"{plan_name}.xlsx"
                workbook.save(os.path.join(folder, excel_filename))
            except Exception as e:
                messagebox.showerror("오류", f"파일 처리 중 오류 발생: {filename}\n{str(e)}")
    messagebox.showinfo("완료", "파일 처리가 완료되었습니다.")

def exit_program():
    root.destroy()

root = tk.Tk()
root.title("Dicom RT Plan 파일 불러오기")

folder_path = tk.StringVar()

frame = tk.Frame(root)
frame.pack(pady=20)

load_button = tk.Button(frame, text="불러오기", command=load_folder)
load_button.pack(side=tk.LEFT, padx=10)

exit_button = tk.Button(frame, text="프로그램 종료", command=exit_program)
exit_button.pack(side=tk.RIGHT, padx=10)

folder_label = tk.Label(root, textvariable=folder_path)
folder_label.pack(pady=10)

root.mainloop()
